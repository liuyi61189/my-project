# -*- coding: utf-8 -*-
"""
墨刀技能 - 5 阶段工作流编排器（后端落地版）

忠实复用现成技能包 modao_skill/ 的 prompts/workflow，映射到 TestHub 后端：
  阶段0 墨刀读取   -> ModaoExtractor（Playwright 逐页 + 注释 + 截图 + 视觉OCR）
  阶段1 需求通读   -> p0_requirement_struct（YAML 结构化摘要）
  阶段2 需求澄清   -> p0 复用（识别模糊/缺失/边界 -> 澄清清单）
  阶段3 用例设计   -> p1拆分 / p2测试点 / p5风险 / p6 PCI / p7三路合并(向量去重) /
                      p3用例 / p4冒烟 / p8质量自检 / Excel导出
  阶段4 冒烟执行   -> 由前端驱动用户决策，后端记录 passed/rejected

向量去重（方案 A）：p7 之前用 Chroma(bge-m3) 对三路测试点做语义近邻去重。
"""
import os
import json
import uuid
import logging
import asyncio
import datetime
import re
import urllib.parse

import numpy as np

from apps.requirement_analysis.modao_prompts import (
    render_prompt,
    P0_REQUIREMENT_STRUCT, P1_MODULE_SPLIT, P2_TESTPOINT_GEN, P3_CASE_GEN,
    P4_SMOKE_EXTRACT, P5_RISK_IDENTIFY, P6_PCI_IDENTIFY, P7_MERGE_DEDUP, P8_QUALITY_CHECK,
)
from apps.requirement_analysis.vector_store import (
    get_merge_collection, release_merge_collection,
)
from apps.requirement_analysis.models import AIModelService

logger = logging.getLogger(__name__)

MEDIA_ROOT = os.environ.get('MEDIA_ROOT', 'media')
DEDUP_THRESHOLD = 0.85  # 余弦相似度阈值（中文需求语义去重）


# ---------- 解析工具 ----------
def _extract_block(content: str, lang: str = None) -> str:
    """从 LLM 输出中提取首个 ```lang 代码块。宽松兼容：
    - 标记后不紧跟换行（如 ` ```json{...} `）
    - 缺少结束 ``` 标记（截至文末）
    - 无代码块时返回原文（供 _parse_json/_parse_yaml 直接解析裸 JSON/YAML）
    """
    if not content:
        return ''
    import re
    # 结束定界符：优先 ```，否则匹配到字符串末尾（兼容未闭合代码块）
    end = r'(?:```|$)'
    if lang:
        pat = re.compile(r'```' + lang + r'\s*\n?' + r'(.*?)' + end, re.DOTALL | re.IGNORECASE)
        m = pat.search(content)
        if m:
            return m.group(1).strip()
    # 通用：匹配 ```yaml / ```json / ```yml 或无语言标记的 ``` 代码块
    pat = re.compile(r'```(?:yaml|json|yml)?\s*\n?' + r'(.*?)' + end, re.DOTALL | re.IGNORECASE)
    m = pat.search(content)
    if m:
        return m.group(1).strip()
    return content.strip()


def _parse_yaml(text: str):
    try:
        import yaml
        return yaml.safe_load(text)
    except Exception as e:
        logger.warning(f'YAML 解析失败: {e}')
        return None


def _parse_json(text: str):
    try:
        return json.loads(text)
    except Exception:
        # 容错：尝试提取首个 { ... } 块
        import re
        m = re.search(r'\{.*\}', text, re.DOTALL)
        if m:
            try:
                return json.loads(m.group(0))
            except Exception:
                pass
        return None

def _normalize_evidence_text(value) -> str:
    """Normalize text for conservative evidence matching."""
    text = str(value or '').lower()
    return re.sub(r'[\s，。；：、,.!?！？:;（）()【】\[\]"“”\'‘’\-—_]+', '', text)


def _evidence_quotes(value):
    """Yield quote strings from legacy strings or structured evidence lists."""
    if isinstance(value, str):
        yield value
    elif isinstance(value, dict):
        quote = value.get('quote') or value.get('evidence') or value.get('text')
        if quote:
            yield str(quote)
    elif isinstance(value, list):
        for item in value:
            yield from _evidence_quotes(item)

def _evidence_items(value):
    """Keep structured page/quote evidence intact while merging modules."""
    if value is None:
        return []
    if isinstance(value, list):
        return list(value)
    return [value]


def _validate_evidence(value, source_text: str) -> bool:
    """Evidence is valid only when every meaningful quote occurs in source."""
    source = _normalize_evidence_text(source_text)
    quotes = [
        _normalize_evidence_text(q)
        for q in _evidence_quotes(value)
        if q and '未找到直接原文' not in str(q)
    ]
    # Short fragments are too weak to be trustworthy evidence.
    if not quotes or any(len(q) < 8 for q in quotes):
        return False
    return all(q in source for q in quotes)


def _ordered_union(left, right):
    result = []
    for item in list(left or []) + list(right or []):
        if item not in result:
            result.append(item)
    return result


def _module_signature(module):
    parts = [
        module.get('name', ''),
        module.get('description', ''),
        *[str(x) for x in module.get('key_features', []) if x],
    ]
    text = _normalize_evidence_text(''.join(parts))
    for word in ('模块', '功能', '新增', '修改', '优化', '修订', '补充', '逻辑', '页面'):
        text = text.replace(word, '')
    return {text[i:i + 2] for i in range(max(0, len(text) - 1))}


def _overlap_coefficient(left, right) -> float:
    if not left or not right:
        return 0.0
    return len(left & right) / min(len(left), len(right))


def _should_merge_modules(left, right) -> bool:
    left_name = _normalize_evidence_text(left.get('name'))
    right_name = _normalize_evidence_text(right.get('name'))
    generic_words = ('模块', '功能', '新增', '修改', '优化', '修订', '补充', '逻辑', '页面')
    for word in generic_words:
        left_name = left_name.replace(word, '')
        right_name = right_name.replace(word, '')
    if (len(left_name) >= 2 and len(right_name) >= 2
            and (left_name in right_name or right_name in left_name)):
        return True

    left_sources = {str(x) for x in left.get('source_locations', []) if x}
    right_sources = {str(x) for x in right.get('source_locations', []) if x}
    source_overlap = _overlap_coefficient(left_sources, right_sources)
    content_overlap = _overlap_coefficient(
        _module_signature(left), _module_signature(right)
    )
    return source_overlap >= 0.5 and content_overlap >= 0.45


def _merge_module_pair(primary, duplicate):
    merged = dict(primary)
    merged['merged_from'] = _ordered_union(
        primary.get('merged_from') or [primary.get('name')],
        duplicate.get('merged_from') or [duplicate.get('name')],
    )
    merged['source_locations'] = _ordered_union(
        primary.get('source_locations'), duplicate.get('source_locations')
    )
    merged['key_features'] = _ordered_union(
        primary.get('key_features'), duplicate.get('key_features')
    )

    merged['source_evidence'] = _ordered_union(
        _evidence_items(primary.get('source_evidence')),
        _evidence_items(duplicate.get('source_evidence')),
    )

    priority_rank = {'P0': 0, 'P1': 1, 'P2': 2}
    if priority_rank.get(duplicate.get('priority'), 9) < priority_rank.get(
        primary.get('priority'), 9
    ):
        merged['priority'] = duplicate.get('priority')
        merged['priority_source'] = duplicate.get('priority_source')
        merged['priority_reason'] = duplicate.get('priority_reason')

    for key in ('business_flow',):
        left_value = str(primary.get(key) or '').strip()
        right_value = str(duplicate.get(key) or '').strip()
        if right_value and right_value not in left_value:
            merged[key] = '\n'.join(x for x in (left_value, right_value) if x)
    return merged


def _postprocess_requirement_summary(data, source_text: str):
    """Merge probable duplicate modules and annotate evidence validity."""
    if not isinstance(data, dict):
        return data, 0, 0

    merged_modules = []
    merge_count = 0
    for raw_module in data.get('modules') or []:
        if not isinstance(raw_module, dict):
            continue
        module = dict(raw_module)
        for index, existing in enumerate(merged_modules):
            if _should_merge_modules(existing, module):
                merged_modules[index] = _merge_module_pair(existing, module)
                merge_count += 1
                break
        else:
            merged_modules.append(module)

    invalid_count = 0
    checked_count = 0
    priority_is_explicit = bool(re.search(r'(?<![A-Za-z0-9])P[012](?![A-Za-z0-9])', source_text, re.I))
    for module in merged_modules:
        if not priority_is_explicit:
            module['priority_source'] = '模型建议'
        module['business_flow_source'] = (
            '原文明示' if _validate_evidence(module.get('business_flow'), source_text)
            else '模型推演'
        )
        valid = _validate_evidence(module.get('source_evidence'), source_text)
        module['evidence_valid'] = valid
        module['evidence_level'] = 'A' if valid else 'D'
        module['human_confirmation_required'] = not valid
        checked_count += 1
        if not valid:
            invalid_count += 1
        for key in ('ambiguous_points', 'boundary_points'):
            for item in module.get(key) or []:
                if not isinstance(item, dict):
                    continue
                valid = _validate_evidence(item.get('evidence'), source_text)
                item['evidence_valid'] = valid
                item['evidence_level'] = 'A' if valid else 'D'
                item['human_confirmation_required'] = True
                checked_count += 1
                invalid_count += int(not valid)

    for key in ('ambiguous_points', 'boundary_points'):
        for item in data.get(key) or []:
            if isinstance(item, dict):
                valid = _validate_evidence(item.get('evidence'), source_text)
                item['evidence_valid'] = valid
                item['evidence_level'] = 'A' if valid else 'D'
                item['human_confirmation_required'] = True
                checked_count += 1
                if not valid:
                    invalid_count += 1

    data['modules'] = merged_modules
    data['evidence_validation'] = {
        'checked_count': checked_count,
        'invalid_count': invalid_count,
        'rule': 'normalized_exact_quote_match',
        'levels': {
            'A': '原文或视觉元素精确验证，可进入测试设计',
            'B': '跨页面推导，需保留推导标记',
            'C': '模型建议或需求缺失，仅用于探索性分析',
            'D': '证据验证失败，必须人工确认后才能生成最终用例',
        },
    }
    for item in data.get('missing_points') or []:
        if isinstance(item, dict):
            item['evidence_level'] = 'C'
            item['human_confirmation_required'] = True
    return data, merge_count, invalid_count


class ModaoWorkflow:
    """墨刀技能工作流编排。"""

    def __init__(self, prototype, analyzer_config, writer_config, embedder_config,
                 vision_config=None, dedup_threshold=DEDUP_THRESHOLD):
        self.prototype = prototype
        self.analyzer_config = analyzer_config
        self.writer_config = writer_config
        self.embedder_config = embedder_config
        self.vision_config = vision_config
        self.dedup_threshold = dedup_threshold
        self.session_id = prototype.uuid

    # ---------- 配置解析 ----------
    @staticmethod
    def resolve_configs(analyzer_id=None, writer_id=None, embedder_id=None):
        from apps.requirement_analysis.models import AIModelConfig
        analyzer = (AIModelConfig.objects.filter(id=analyzer_id).first()
                    if analyzer_id else AIModelConfig.get_active_config('siliconflow', 'analyzer')
                    or AIModelConfig.objects.filter(role='analyzer', is_active=True).first())
        writer = (AIModelConfig.objects.filter(id=writer_id).first()
                  if writer_id else AIModelConfig.objects.filter(role='writer', is_active=True).first())
        embedder = (AIModelConfig.objects.filter(id=embedder_id).first()
                    if embedder_id else AIModelConfig.get_active_embedder())
        # fallback: analyzer 缺失时复用 writer（两者都是 LLM 文本生成，可共用模型）
        if not analyzer and writer:
            analyzer = writer
        return analyzer, writer, embedder

    # ---------- LLM / 向量 基础 ----------
    def _llm(self, config, prompt: str, max_tokens: int = 8000, timeout: int = 300):
        """调用 LLM。带超时保护：单条调用超过 timeout 秒即抛 asyncio.TimeoutError，
        由上层（run_design_phase1 模块级 try/except）捕获并降级，避免窗口永久卡死。"""
        if not config:
            raise Exception('缺少可用的 AI 模型配置（analyzer/writer）')
        messages = [{'role': 'user', 'content': prompt}]
        try:
            result = asyncio.run(
                asyncio.wait_for(
                    AIModelService.call_openai_compatible_api(config, messages, max_tokens=max_tokens),
                    timeout=timeout,
                )
            )
        except asyncio.TimeoutError:
            raise asyncio.TimeoutError(
                f'LLM 调用超时（>{timeout}s），请检查模型服务是否可用或网络是否通畅'
            )
        return result['choices'][0]['message']['content']

    def _embed(self, texts):
        if not self.embedder_config:
            return None
        try:
            return asyncio.run(
                asyncio.wait_for(
                    AIModelService.embed_texts(self.embedder_config, texts),
                    timeout=300,
                )
            )
        except asyncio.TimeoutError:
            logger.warning(f'向量化超时（降级为无去重，>300s）')
            return None
        except Exception as e:
            logger.warning(f'向量化失败（降级为无去重）: {e}')
            return None

    # ---------- 向量语义去重（方案 A）----------
    def _vector_dedup(self, items):
        """
        items: [{'id','title','description','source'(list),'priority'}]
        返回 (deduped_items, dedup_count)
        """
        if not items:
            return items, 0
        texts = [it.get('description') or it.get('title') or '' for it in items]
        vecs = self._embed(texts)
        if vecs is None:
            logger.info('无 embedder，跳过向量去重')
            return items, 0
        try:
            arr = np.array(vecs, dtype=float)
            norms = np.linalg.norm(arr, axis=1, keepdims=True)
            arr = arr / np.clip(norms, 1e-9, None)
            sim = arr @ arr.T
        except Exception as e:
            logger.warning(f'向量去重计算失败: {e}')
            return items, 0

        n = len(items)
        keep = [True] * n
        dedup = 0
        # 优先级权重：P0>P1>P2
        prio_w = {'P0': 0, 'P1': 1, 'P2': 2}

        def better(a, b):
            return prio_w.get(a.get('priority'), 9) <= prio_w.get(b.get('priority'), 9)

        for i in range(n):
            if not keep[i]:
                continue
            for j in range(i + 1, n):
                if keep[j] and sim[i][j] >= self.dedup_threshold:
                    # 合并来源标记，保留高优先级代表
                    src = list(set((items[i].get('source') or []) + (items[j].get('source') or [])))
                    if better(items[j], items[i]):
                        items[i]['priority'] = items[j]['priority']
                    items[i]['source'] = src
                    items[i]['description'] = items[i].get('description') or items[j].get('description')
                    keep[j] = False
                    dedup += 1
        result = [items[i] for i in range(n) if keep[i]]
        return result, dedup

    def _log(self, msg: str):
        """追加工作日志到 prototype.work_log（带时间戳）。"""
        from datetime import datetime as dt
        ts = dt.now().strftime('%H:%M:%S')
        entry = f"[{ts}] {msg}\n"
        self.prototype.work_log = (self.prototype.work_log or '') + entry
        self.prototype.save(update_fields=['work_log'])

    def ask_question(self, stage: int, question: str, max_tokens: int = 4000) -> str:
        """针对某阶段产物进行答疑（人工疑点解答），返回 AI 回答文本。"""
        ctx_map = {
            0: (self.prototype.extracted_json or '') + '\n' + (self.prototype.requirement_summary or ''),
            1: self.prototype.requirement_summary or '',
            2: self.prototype.clarification_log or '',
            3: '\n'.join([
                self.prototype.testcases_json or '',
                self.prototype.final_testpoints_json or '',
                self.prototype.risks_json or '',
                self.prototype.pci_json or '',
            ]),
        }
        context = (ctx_map.get(stage, self.prototype.requirement_summary) or '').strip()
        if not context:
            return '该阶段暂无内容可答疑，请先运行对应阶段生成产物。'
        prompt = (
            '你是一名资深测试分析师。下面是「需求读取与用例生成」工作流'
            f'第{stage}阶段的产物内容：\n\n'
            f'===== 阶段{stage}内容 =====\n{context}\n===== 内容结束 =====\n\n'
            f'用户疑问：{question}\n\n'
            '请基于上述内容专业、准确地回答用户疑问；若内容存在错误、遗漏或歧义，请明确指出并给出改进建议。'
            '回答用中文，条理清晰。'
        )
        return self._llm(self.analyzer_config or self.writer_config, prompt, max_tokens=max_tokens)

    # ---------- 阶段0：原型读取（支持多地址批量 + 单地址自动发现页面树）----------
    @staticmethod
    def _page_label_from_url(u: str) -> str:
        """从 URL 推断页面名：优先取 hash 中的 p= 参数（已 URL 解码），否则取路径末段。"""
        try:
            frag = u.split('#', 1)[1] if '#' in u else ''
            m = re.search(r'[?&]p=([^&]+)', frag)
            if not m:
                m = re.search(r'[?&]p=([^&]+)', u)
            if m:
                return urllib.parse.unquote(m.group(1))
        except Exception:
            pass
        path = u.split('#')[0].rstrip('/')
        return path.split('/')[-1] or u

    def run_extract(self, urls=None, url=None, cookies=None, headless=True, source_type=None):
        from apps.requirement_analysis.modao_extractor import ModaoExtractor
        source_type = source_type or self.prototype.source_type or 'modao'
        # 规范化 URL 列表（兼容单字符串 / 列表 / 换行或逗号分隔）
        if urls is None:
            urls = [url] if url else []
        elif isinstance(urls, str):
            urls = [x for x in urls.replace(',', '\n').split('\n') if x.strip()]
        urls = [x.strip() for x in urls if isinstance(x, str) and x.strip()]
        urls = list(dict.fromkeys(urls))  # 去重保序
        if not urls:
            self.prototype.status = 'failed'
            self.prototype.error_message = '未提供有效的需求地址'
            self.prototype.save(update_fields=['status', 'error_message'])
            self._log('未提供有效的需求地址，读取中止')
            return {'pages': []}

        self.prototype.status = 'extracting'
        self.prototype.save(update_fields=['status'])
        self._log(f'开始需求读取，共 {len(urls)} 个地址...')
        extractor = ModaoExtractor(run_id=self.session_id)
        all_pages = []
        for ui, u in enumerate(urls, 1):
            try:
                # webshare：始终枚举页面树（目录地址自动发现所有子页面）。
                # 墨刀多地址：每个当单页抓（避免页面树枚举导致跨地址重复）；
                # 墨刀单地址：尝试枚举页面树自动发现目录内多页。
                if source_type == 'webshare':
                    enumerate_pages = True
                    single = False
                else:
                    single = len(urls) > 1
                    enumerate_pages = not single
                data = extractor.extract_sync(url=u, cookies=cookies, headless=headless,
                                              enumerate_pages=enumerate_pages,
                                              source_type=source_type,
                                              vision_config=self.vision_config)
                for p in data['pages']:
                    if single:
                        # 单页模式 extractor 给的是默认"第N页"，用 URL 推断更可读的页面名
                        p['page_name'] = self._page_label_from_url(u) or p.get('page_name') or f'地址{ui}'
                    all_pages.append(p)
                self._log(f'地址{ui}读取完成：{len(data["pages"])} 页')
            except Exception as e:
                logger.error(f'地址[{u}]读取失败: {e}')
                self._log(f'地址{ui}读取失败: {e}')
                continue

        # 重新编号并落库（先清空旧 pages，避免重跑累积）
        for idx, p in enumerate(all_pages, 1):
            p['page_no'] = idx
        self.prototype.pages.all().delete()
        for p in all_pages:
            self.prototype.pages.create(
                page_no=p['page_no'],
                page_name=p['page_name'],
                text=p.get('text', ''),
                annotations=json.dumps(p.get('annotations', []), ensure_ascii=False),
                screenshot=p.get('screenshot', ''),
            )
        self._log(f'全部读取完成，共 {len(all_pages)} 页')
        self.prototype.extracted_json = json.dumps({
            'urls': urls,
            'pages': [{
                'page_no': p['page_no'],
                'page_name': p['page_name'],
                'text': p.get('text', ''),
                'annotations': p.get('annotations', []),
                'screenshot': p.get('screenshot', ''),
                'visual_summary': p.get('visual_summary', ''),
                'visual_artifacts': p.get('visual_artifacts', []),
            } for p in all_pages],
            'warnings': [],
        }, ensure_ascii=False)
        self.prototype.source_type = source_type
        self.prototype.status = 'extracted'
        self.prototype.current_stage = 0
        self.prototype.save()
        return {'pages': all_pages}

    # ---------- 阶段1：需求结构化 ----------
    def _source_text(self) -> str:
        if self.prototype.extracted_json:
            data = json.loads(self.prototype.extracted_json)
            parts = []
            for p in data.get('pages', []):
                parts.append(
                    f"### {p['page_name']}\n{p.get('text','')}\n"
                    f"[注释/标注] {json.dumps(p.get('annotations', []), ensure_ascii=False)}\n"
                    f"[结构化视觉内容 visual_artifacts] "
                    f"{json.dumps(p.get('visual_artifacts', []), ensure_ascii=False)}")
            return '\n\n'.join(parts)
        return self.prototype.requirement_summary or ''

    def _source_decision_tables(self):
        """Return decision tables from vision extraction as authoritative source data."""
        if not self.prototype.extracted_json:
            return []
        try:
            data = json.loads(self.prototype.extracted_json)
        except (TypeError, ValueError):
            return []
        tables = []
        for page in data.get('pages') or []:
            for artifact in page.get('visual_artifacts') or []:
                if not isinstance(artifact, dict) or artifact.get('artifact_type') != 'decision_table':
                    continue
                table = json.loads(json.dumps(artifact, ensure_ascii=False))
                table.setdefault('source_page', f"第{page.get('page_no')}页")
                table.setdefault('name', table.get('title') or table.get('artifact_id') or '决策表')
                tables.append(table)
        return tables

    def _source_visual_structures(self):
        """Return all non-trivial visual structures with page traceability."""
        if not self.prototype.extracted_json:
            return []
        try:
            data = json.loads(self.prototype.extracted_json)
        except (TypeError, ValueError):
            return []
        trivial = {'ui', 'annotation'}
        structures = []
        for page in data.get('pages') or []:
            for artifact in page.get('visual_artifacts') or []:
                if not isinstance(artifact, dict) or artifact.get('artifact_type') in trivial:
                    continue
                item = json.loads(json.dumps(artifact, ensure_ascii=False))
                item.setdefault('source_page', f"第{page.get('page_no')}页")
                item.setdefault('source_artifact_id', item.get('artifact_id'))
                structures.append(item)
        return structures

    def _restore_authoritative_decision_tables(self, parsed):
        """Keep LLM module association, but never trust it to rewrite visual matrix cells."""
        source_tables = self._source_decision_tables()
        if not isinstance(parsed, dict) or not source_tables:
            return parsed
        generated = {
            item.get('source_artifact_id') or item.get('artifact_id'): item
            for item in (parsed.get('decision_tables') or []) if isinstance(item, dict)
        }
        restored = []
        for source in source_tables:
            artifact_id = source.get('artifact_id')
            enriched = generated.get(artifact_id) or {}
            table = dict(source)
            table['source_artifact_id'] = artifact_id
            if enriched.get('related_module'):
                table['related_module'] = enriched['related_module']
            restored.append(table)
        parsed['decision_tables'] = restored
        return parsed

    def _restore_authoritative_visual_structures(self, parsed):
        """Preserve source structures; LLM may only enrich their module association/name."""
        if not isinstance(parsed, dict):
            return parsed
        source_items = self._source_visual_structures()
        generated_items = []
        for key in ('visual_structure_links', 'visual_structures', 'state_matrices', 'decision_tables'):
            generated_items.extend(parsed.get(key) or [])
        generated = {
            item.get('source_artifact_id') or item.get('artifact_id'): item
            for item in generated_items if isinstance(item, dict)
        }
        modules = [item for item in (parsed.get('modules') or []) if isinstance(item, dict)]
        restored = []
        for source in source_items:
            artifact_id = source.get('artifact_id')
            enrichment = generated.get(artifact_id) or {}
            item = dict(source)
            item['source_artifact_id'] = artifact_id
            if enrichment.get('name'):
                item['name'] = enrichment['name']
            source_page = str(item.get('source_page') or '')
            page_candidates = [
                module.get('name') for module in modules
                if source_page and source_page in [str(x) for x in (module.get('source_locations') or [])]
            ]
            proposed = enrichment.get('related_module')
            if proposed in page_candidates:
                item['related_module'] = proposed
            elif len(page_candidates) == 1:
                item['related_module'] = page_candidates[0]
            elif proposed:
                item['related_module'] = proposed
                item['module_association_warning'] = '所属模块未通过来源页交叉校验'
            restored.append(item)
        parsed['visual_structures'] = restored
        parsed['state_matrices'] = [
            item for item in restored if item.get('artifact_type') == 'state_matrix'
        ]
        return parsed

    # ---------- 项目知识库（辅助生成 + 自适应闭环）----------
    def _kb_context(self) -> str:
        """读取本项目知识库上下文，用于辅助 AI 生成；无项目/无知识库返回空串（降级不阻塞）。"""
        if getattr(self, '_kb_cache', None) is not None:
            return self._kb_cache
        ctx = ''
        try:
            if self.prototype.project_id:
                from apps.projects.models import Project
                from apps.requirement_analysis.models import AIModelService
                project = Project.objects.filter(id=self.prototype.project_id).first()
                if project:
                    ctx = AIModelService.build_knowledge_context_from_project(project) or ''
        except Exception as e:
            logger.warning(f'读取项目知识库失败（降级为无知识库）: {e}')
            ctx = ''
        self._kb_cache = ctx
        return ctx

    def _kb_suffix(self) -> str:
        """拼接到各阶段 prompt 末尾的知识库参考段（无知识库时返回空串）。"""
        ctx = self._kb_context()
        if not ctx:
            return ''
        return (
            "\n\n### 项目知识库参考\n"
            + ctx
            + "\n（请结合上述项目业务背景与既有测试约束来辅助本次分析与生成；"
              "不要重复知识库中已明确的内容，聚焦于本次需求新增/变化的部分）"
        )

    def run_struct(self):
        self.prototype.status = 'structuring'
        self.prototype.save(update_fields=['status'])
        self._log('开始需求结构化分析...')
        text = self._source_text()
        prompt = render_prompt(P0_REQUIREMENT_STRUCT, requirement_text=text) + self._kb_suffix()
        out = self._llm(self.analyzer_config, prompt, max_tokens=8000)
        parsed = _parse_yaml(_extract_block(out, 'yaml'))
        if isinstance(parsed, dict):
            parsed = self._restore_authoritative_decision_tables(parsed)
            parsed = self._restore_authoritative_visual_structures(parsed)
            parsed, merge_count, invalid_count = _postprocess_requirement_summary(
                parsed, text
            )
            import yaml
            out = yaml.safe_dump(
                parsed, allow_unicode=True, sort_keys=False, width=120
            )
            if merge_count:
                self._log(f'结构化后处理：合并 {merge_count} 个重复模块')
            if invalid_count:
                self._log(f'证据校验：{invalid_count} 项未匹配原始页面，已标记 evidence_valid=false')
        self.prototype.requirement_summary = out
        self.prototype.status = 'extracted'
        self.prototype.current_stage = 1
        self.prototype.save()
        self._log('需求结构化完成，已生成摘要')
        return out

    # ---------- 阶段2：需求澄清 ----------
    def run_clarify(self):
        self.prototype.status = 'clarifying'
        self.prototype.save(update_fields=['status'])
        self._log('开始需求澄清（识别模糊/缺失/边界）...')
        # 阶段1摘要决定已确认范围；原始内容只用于证据核对和查漏，
        # 避免摘要遗漏被继续放大，也避免把用户已排除的需求重新带回流程。
        structured = (self.prototype.requirement_summary or '').strip()
        source_text = self._source_text().strip()
        confirmed_scope = structured or source_text
        scope_note = (
            '结构化摘要是本阶段的需求范围边界。原始内容仅用于验证摘要是否有遗漏或无依据表述；'
            '不要针对摘要中已经排除、标记不做或超出范围的功能提问。'
            if structured and source_text else
            '当前没有独立的结构化摘要，请直接基于原始内容审查。'
        )
        prompt = (
            '你是一名严格、保守的需求审查员。你的任务不是补写需求，而是对照原始内容审查'
            '已确认范围中的模糊点、完成当前流程所必需但缺失的信息，以及功能边界。\n\n'
            f'===== 已确认的结构化范围 =====\n{confirmed_scope}\n===== 范围结束 =====\n\n'
            f'===== 原始原型/需求内容（仅用于证据核对与查漏） =====\n'
            f'{source_text}\n===== 原始内容结束 =====\n\n'
            f'### 范围规则\n{scope_note}\n\n'
            '### 审查规则\n'
            '1. 每个问题必须能指向具体页面、标注、字段或结构化条目。\n'
            '2. 原文明确写出的内容不是缺失点；不得重复追问。\n'
            '3. 只有完成当前已确认流程确实必需的信息，才能判为缺失点。\n'
            '4. 如果结构化摘要遗漏了范围内的原文，audit_kind 写 "结构化遗漏"。\n'
            '5. 如果结构化摘要中的表述在原文找不到依据，audit_kind 写 "结构化无依据"。\n'
            '6. 不得为了凑数量编造问题；没有可靠问题时输出空数组。\n\n'
            '### 输出要求\n'
            '**必须且仅**输出一个 JSON 数组，不要使用 markdown 代码块或附加说明。'
            '数组每项字段如下：\n'
            '- type: 仅限 "模糊点" / "缺失点" / "边界点"\n'
            '- audit_kind: 仅限 "原文歧义" / "结构化遗漏" / "结构化无依据" / "信息缺失" / "边界不清"\n'
            '- location: 具体页面、标注、字段或结构化条目\n'
            '- evidence: 支持判断的简短原文；找不到直接依据时写 "未找到直接原文"\n'
            '- issue: 客观描述问题，不得把推断写成事实\n'
            '- suggested_question: 可由产品明确回答的单一问题\n'
            '- priority: "高" / "中" / "低"\n'
            '- impact: 受影响的测试场景或模块\n\n'
            '示例（仅表示格式，不得复制内容）：\n'
            '[{"type":"缺失点","audit_kind":"信息缺失","location":"登录页/密码字段",'
            '"evidence":"原文仅写登录密码","issue":"密码规则未说明",'
            '"suggested_question":"密码允许的长度和字符范围是什么？","priority":"中","impact":"登录校验"}]'
        )
        out = self._llm(self.analyzer_config, prompt, max_tokens=6000)
        try:
            clarify_items = json.loads(out)
        except Exception:
            clarify_items = None
        if isinstance(clarify_items, list):
            invalid_count = 0
            for item in clarify_items:
                if not isinstance(item, dict):
                    continue
                valid = _validate_evidence(item.get('evidence'), source_text)
                item['evidence_valid'] = valid
                invalid_count += int(not valid)
            out = json.dumps(clarify_items, ensure_ascii=False, indent=2)
            if invalid_count:
                self._log(f'澄清证据校验：{invalid_count} 项未匹配原始页面')
        self.prototype.clarification_log = out
        self.prototype.status = 'extracted'
        self.prototype.current_stage = 2
        self.prototype.save()
        self._log('需求澄清完成')
        return out

    # ---------- 阶段3：用例设计（核心）----------
    def run_design(self):
        """完整跑阶段3（兼容旧入口）：第一段测试点合并 + 第二段用例生成，一口气到底。"""
        self.run_design_phase1()
        return self.run_design_phase2()

    # ---------- 阶段3 第一段：模块拆分→测试点→风险→PCI→三路合并 ----------
    def run_design_phase1(self):
        """产出 module_split / risks_json / pci_json / final_testpoints_json，
        完成后 status='testpoints_ready'，等待人工确认/编辑测试点后再调 phase2 生成用例。"""
        self.prototype.status = 'designing'
        self.prototype.save(update_fields=['status'])
        self._log('开始用例设计（阶段3 · 第一段：测试点合并）...')

        # 3.1 模块拆分
        self._log('3.1 正在拆分模块...')
        split_prompt = render_prompt(P1_MODULE_SPLIT, requirement_yaml=self.prototype.requirement_summary) + self._kb_suffix()
        split_out = self._llm(self.analyzer_config, split_prompt, max_tokens=6000)
        split_yaml = _parse_yaml(_extract_block(split_out, 'yaml')) or {}
        modules = split_yaml.get('modules', []) if isinstance(split_yaml, dict) else []
        summary_yaml = _parse_yaml(_extract_block(self.prototype.requirement_summary or '', 'yaml')) or {}
        decision_tables = (
            summary_yaml.get('decision_tables', [])
            if isinstance(summary_yaml, dict) else []
        ) or self._source_decision_tables()
        visual_structures = (
            summary_yaml.get('visual_structures', [])
            if isinstance(summary_yaml, dict) else []
        ) or self._source_visual_structures()
        for table in decision_tables:
            related = _normalize_evidence_text(table.get('related_module'))
            candidates = [
                module for module in modules
                if related and (
                    related in _normalize_evidence_text(module.get('name'))
                    or _normalize_evidence_text(module.get('name')) in related
                )
            ]
            if not candidates and len(modules) == 1:
                candidates = modules
            for module in candidates:
                module.setdefault('decision_tables', []).append(table)
        for structure in visual_structures:
            if structure.get('artifact_type') == 'decision_table':
                continue
            related = _normalize_evidence_text(structure.get('related_module'))
            candidates = [
                module for module in modules
                if related and (
                    related in _normalize_evidence_text(module.get('name'))
                    or _normalize_evidence_text(module.get('name')) in related
                )
            ]
            if not candidates and len(modules) == 1:
                candidates = modules
            for module in candidates:
                module.setdefault('visual_structures', []).append(structure)
        if isinstance(split_yaml, dict):
            import yaml
            split_out = yaml.safe_dump(split_yaml, allow_unicode=True, sort_keys=False, width=120)
        self.prototype.module_split = split_out
        self._log(f'3.1 模块拆分完成，共 {len(modules)} 个模块')

        all_final_tps = []      # 全部模块的最终测试点
        all_risks = []
        all_pci = []

        for idx, mod in enumerate(modules, 1):
            mod_name = mod.get('name', '未命名模块')
            self._log(f'3.2 开始处理模块[{idx}/{len(modules)}] {mod_name} ...')
            mod_ctx = json.dumps(mod, ensure_ascii=False)
            try:
                # 3.3 基础测试点 (draft)
                tp_prompt = render_prompt(P2_TESTPOINT_GEN, module_info_yaml=mod_ctx)
                tp_out = self._llm(self.analyzer_config, tp_prompt, max_tokens=6000)
                tp_yaml = _parse_yaml(_extract_block(tp_out, 'yaml')) or {}
                draft_tps = tp_yaml.get('test_points', []) if isinstance(tp_yaml, dict) else []
                for t in draft_tps:
                    t['source'] = ['draft']

                # 3.3.5 风险识别 (risk)
                risk_prompt = render_prompt(P5_RISK_IDENTIFY,
                                            test_points_json=json.dumps(draft_tps, ensure_ascii=False))
                risk_out = self._llm(self.analyzer_config, risk_prompt, max_tokens=6000)
                risk_json = _parse_json(_extract_block(risk_out, 'json')) or {}
                all_risks.append({'module': mod_name, 'data': risk_json})
                risk_ext = []
                for rp in risk_json.get('risk_points', []):
                    for etp in rp.get('extended_test_points', []):
                        risk_ext.append({
                            'id': etp.get('id', f"{mod_name}-EXT"),
                            'title': etp.get('title', ''),
                            'description': etp.get('description', ''),
                            'priority': etp.get('priority_hint', 'P1'),
                            'source': ['risk'],
                        })

                # 3.3.6 PCI 识别 (pci)
                pci_prompt = render_prompt(P6_PCI_IDENTIFY,
                                           test_points_json=json.dumps(draft_tps, ensure_ascii=False),
                                           requirement_text=self._source_text(),
                                           stage2_issues=self.prototype.clarification_log)
                pci_out = self._llm(self.analyzer_config, pci_prompt, max_tokens=6000)
                pci_json = _parse_json(_extract_block(pci_out, 'json')) or {}
                all_pci.append({'module': mod_name, 'data': pci_json})

                # 3.3.7 三路合并去重（向量去重 + p7 最终化）
                combined = []
                for t in (draft_tps + risk_ext):
                    combined.append({
                        'id': t.get('id', f"{mod_name}-{len(combined)}"),
                        'title': t.get('title', ''),
                        'description': t.get('description', ''),
                        'priority': t.get('priority', 'P1'),
                        'source': t.get('source', ['draft']),
                    })
                deduped, dedup_count = self._vector_dedup(combined)
                merge_prompt = render_prompt(
                    P7_MERGE_DEDUP,
                    draft_testpoints=json.dumps(deduped, ensure_ascii=False),
                    risk_extended_testpoints='[]',
                    pci_blocked_scenarios=json.dumps(
                        [p for p in pci_json.get('pci_list', [])], ensure_ascii=False),
                )
                merge_out = self._llm(self.analyzer_config, merge_prompt, max_tokens=8000)
                merge_json = _parse_json(_extract_block(merge_out, 'json')) or {}
                # 兼容部分模型返回 YAML 而非 JSON 的情况
                if not merge_json:
                    merge_json = _parse_yaml(_extract_block(merge_out, 'yaml')) or {}
                final_tps = merge_json.get('test_points', []) if isinstance(merge_json, dict) else []
                # 安全网：若 P7 未按约定返回测试点，但去重后输入非空，则降级使用输入
                if not final_tps and deduped:
                    logger.warning(f'模块[{mod_name}] P7 返回空测试点，降级使用去重后输入({len(deduped)}条)')
                    self._log(f'  模块[{mod_name}] P7 返回空，降级使用去重后输入 {len(deduped)} 条')
                    final_tps = deduped
                for t in final_tps:
                    t['module'] = mod_name
                all_final_tps.extend(final_tps)
                logger.info(f'模块[{mod_name}] 三路合并：输入{len(combined)} 去重{dedup_count} 最终{len(final_tps)}')
                self._log(f'  模块[{mod_name}] 合并完成：{len(final_tps)} 测试点')
            except Exception as e:
                logger.error(f'模块[{mod_name}] 测试点合并失败: {e}')
                self._log(f'  模块[{mod_name}] 失败: {e}')
                continue

        # P7 合并后全局二次去重：消除各模块合并可能 reintroduce 的近重复
        # 按 module 分组去重，仅消同模块内的近重复，避免跨模块同功能被误删
        global_dedup = 0
        if self.embedder_config and all_final_tps:
            grouped = {}
            for t in all_final_tps:
                grouped.setdefault(t.get('module', '') or '', []).append(t)
            merged = []
            for mod_name, mod_items in grouped.items():
                deduped, cnt = self._vector_dedup(mod_items)
                global_dedup += cnt
                merged.extend(deduped)
            all_final_tps = merged
            if global_dedup:
                self._log(f'  全局二次去重：消除 {global_dedup} 条合并重新引入的近重复')

        # 第一段产物落库（测试点合并完成，等待人工确认）
        self.prototype.risks_json = json.dumps(all_risks, ensure_ascii=False)
        self.prototype.pci_json = json.dumps(all_pci, ensure_ascii=False)
        self.prototype.final_testpoints_json = json.dumps(all_final_tps, ensure_ascii=False)
        self.prototype.status = 'testpoints_ready'
        self.prototype.current_stage = 3
        self.prototype.save()
        self._log(f'第一段完成！共 {len(modules)} 模块 / {len(all_final_tps)} 测试点。请确认/编辑测试点后再生成测试用例。')
        release_merge_collection(self.session_id)
        return {
            'modules': len(modules),
            'final_testpoints': len(all_final_tps),
            'phase': 1,
        }

    # ---------- 阶段3 第二段：基于（已确认的）测试点生成用例→冒烟→质量→Excel ----------
    def run_design_phase2(self):
        """读取（可能被人工编辑过的）final_testpoints_json，生成测试用例、冒烟、质量报告、Excel。"""
        self.prototype.status = 'designing'
        self.prototype.save(update_fields=['status'])
        self._log('开始生成测试用例（阶段3 · 第二段）...')

        # 读取已确认的最终测试点
        try:
            all_final_tps = json.loads(self.prototype.final_testpoints_json or '[]')
            if not isinstance(all_final_tps, list):
                all_final_tps = []
        except Exception:
            all_final_tps = []
        if not all_final_tps:
            self.prototype.status = 'failed'
            self.prototype.error_message = '没有可用的测试点，请先运行第一段生成测试点'
            self.prototype.save(update_fields=['status', 'error_message'])
            self._log('第二段中止：final_testpoints_json 为空')
            return {'testcases': 0, 'error': 'no_testpoints'}

        # 重新解析模块拆分，供用例生成提供 module_context
        split_yaml = _parse_yaml(_extract_block(self.prototype.module_split, 'yaml')) or {}
        modules = split_yaml.get('modules', []) if isinstance(split_yaml, dict) else []
        mod_map = {m.get('name', ''): m for m in modules if isinstance(m, dict)}

        # 按模块分组测试点（保持原始顺序）
        tps_by_mod = {}
        for tp in all_final_tps:
            mod_name = (tp.get('module') if isinstance(tp, dict) else None) or '未分组'
            tps_by_mod.setdefault(mod_name, []).append(tp)

        # 3.4 用例生成（按模块批量）
        module_testcases = []
        for mod_name, final_tps in tps_by_mod.items():
            if not final_tps:
                continue
            mod = mod_map.get(mod_name, {'name': mod_name})
            mod_ctx = json.dumps(mod, ensure_ascii=False)
            try:
                self._log(f'  模块[{mod_name}] 正在生成测试用例（{len(final_tps)} 测试点）...')
                case_prompt = (render_prompt(P3_CASE_GEN,
                                             test_point_yaml=json.dumps(final_tps, ensure_ascii=False),
                                             module_context=mod_ctx)
                               + "\n\n### 附加指令\n请为上面列出的【每一条】最终测试点生成一条结构化测试用例，"
                                 "以 YAML 列表返回，列表项为 `testcase:` 字典（字段同 p3 输出规范）。"
                               + self._kb_suffix())
                # 注入人工已确认的 PCI 答案 + 风险处置结论，让 AI 基于确认结论精准出例
                confirmed_ctx = self._build_confirmed_context()
                if confirmed_ctx:
                    case_prompt += (
                        "\n\n### 已确认的待确认问题答案与风险处置结论（人工已确认，请务必作为用例设计依据）\n"
                        + confirmed_ctx)
                case_out = self._llm(self.writer_config, case_prompt, max_tokens=16000)
                case_yaml = _parse_yaml(_extract_block(case_out, 'yaml')) or {}
                cases = case_yaml.get('testcases', []) if isinstance(case_yaml, dict) else []
                if not cases and isinstance(case_yaml, list):
                    cases = case_yaml
                for c in cases:
                    c['module'] = mod_name
                module_testcases.append((mod_name, cases))
            except Exception as e:
                logger.error(f'模块[{mod_name}] 用例生成失败: {e}')
                self._log(f'  模块[{mod_name}] 用例生成失败: {e}')
                continue

        # 3.5 冒烟提取（p4 技能输出 YAML，必须用 _parse_yaml 解析；用 _parse_json 会失败导致冒烟为空）
        self._log('3.5 正在提取冒烟用例...')
        smoke_prompt = render_prompt(P4_SMOKE_EXTRACT,
                                     all_test_points_yaml=json.dumps(all_final_tps, ensure_ascii=False))
        smoke_out = self._llm(self.analyzer_config, smoke_prompt, max_tokens=6000)
        smoke_json = _parse_yaml(_extract_block(smoke_out, 'yaml')) or {}
        # 兼容模型直接返回数组的情况，统一包装为 {smoke_overview, smoke_testcases, execution_suggestion}
        if isinstance(smoke_json, list):
            smoke_json = {'smoke_testcases': smoke_json}
        self.prototype.smoke_json = json.dumps(smoke_json, ensure_ascii=False)

        # 汇总全部用例
        all_cases = []
        for _, cases in module_testcases:
            all_cases.extend(cases)
        self.prototype.testcases_json = json.dumps(all_cases, ensure_ascii=False)
        self._log(f'3.5 冒烟用例提取完成，共 {len(all_cases)} 条用例')

        # 3.6 质量自检
        self._log('3.6 正在执行质量自检...')
        quality_prompt = render_prompt(P8_QUALITY_CHECK,
                                       generated_testcases_json=json.dumps(all_cases, ensure_ascii=False),
                                       smoke_testcases_json=json.dumps(smoke_json, ensure_ascii=False))
        quality_out = self._llm(self.analyzer_config, quality_prompt, max_tokens=6000)
        quality_json = _parse_json(_extract_block(quality_out, 'json')) or {}
        self.prototype.quality_report_json = json.dumps(quality_json, ensure_ascii=False)
        self._log('3.6 质量自检完成')

        # 3.7 Excel 导出
        self._log('3.7 正在导出 Excel...')
        excel_path = self._export_excel(all_cases, smoke_json)

        self.prototype.excel_path = excel_path or ''
        self.prototype.status = 'done'
        self.prototype.current_stage = 3
        self.prototype.save()
        self._log(f'用例生成全部完成！共 {len(all_final_tps)} 测试点 / {len(all_cases)} 用例 / Excel: {excel_path or "无"}')
        return {
            'final_testpoints': len(all_final_tps),
            'testcases': len(all_cases),
            'smoke': len(smoke_json.get('smoke_testcases', []) if isinstance(smoke_json, dict) else []),
            'excel': excel_path,
            'phase': 2,
        }

    # ---------- 收集人工已确认的 PCI 答案 + 风险处置结论，供用例生成参考 ----------
    def _build_confirmed_context(self):
        """从 pci_json / risks_json 提取人工已确认的答案，拼接为上下文文本。
        用例生成(p3)注入后 AI 可基于确认结论精准出例。兼容多种结构：
        顶层数组 [{module, data:{pci_list/risk_points}}] / 单层 {data:{...}} / 裸数组。
        """
        parts = []
        # 1) PCI 确认答案
        try:
            pci_raw = json.loads(self.prototype.pci_json or '[]')
            pci_list = []
            if isinstance(pci_raw, list):
                for m in pci_raw:
                    lst = None
                    if isinstance(m, dict):
                        lst = (m.get('data', {}) or {}).get('pci_list') or m.get('pci_list')
                    if isinstance(lst, list):
                        pci_list.extend(lst)
            elif isinstance(pci_raw, dict):
                pci_list = (pci_raw.get('data', {}) or {}).get('pci_list') or pci_raw.get('pci_list') or []
            for pci in pci_list:
                if not isinstance(pci, dict):
                    continue
                desc = (pci.get('description') or '').strip()
                if not desc:
                    continue
                lines = []
                subs = pci.get('sub_questions') or []
                if isinstance(subs, list) and subs:
                    for s in subs:
                        q = (s.get('question') or '').strip()
                        a = (s.get('answer') or '').strip()
                        if q and a:
                            lines.append(f'  - Q: {q}\n    A: {a}')
                else:
                    a = (pci.get('answer') or pci.get('resolution_condition') or '').strip()
                    if a:
                        lines.append(f'  - A: {a}')
                if lines:
                    parts.append(f'【待确认问题 {pci.get("id", "")}】{desc}\n' + '\n'.join(lines))
        except Exception:
            pass
        # 2) 风险处置结论
        try:
            risk_raw = json.loads(self.prototype.risks_json or '[]')
            risk_list = []
            if isinstance(risk_raw, list):
                for m in risk_raw:
                    lst = None
                    if isinstance(m, dict):
                        lst = (m.get('data', {}) or {}).get('risk_points') or m.get('risk_points')
                    if isinstance(lst, list):
                        risk_list.extend(lst)
            elif isinstance(risk_raw, dict):
                risk_list = (risk_raw.get('data', {}) or {}).get('risk_points') or risk_raw.get('risk_points') or []
            for r in risk_list:
                if not isinstance(r, dict):
                    continue
                desc = (r.get('description') or r.get('risk_point') or '').strip()
                mit = (r.get('mitigation') or '').strip()
                if desc and mit:
                    parts.append(f'【风险 {r.get("id", "")}】{desc}\n  - 处置: {mit}')
        except Exception:
            pass
        if not parts:
            return ''
        return '\n\n'.join(parts)

    # ---------- Excel 导出（复用技能 module_exporter 样式）----------
    def _export_excel(self, testcases, smoke_json):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except Exception as e:
            logger.warning(f'openpyxl 不可用，跳过 Excel: {e}')
            return ''

        out_dir = os.path.join(MEDIA_ROOT, 'modao', self.session_id)
        os.makedirs(out_dir, exist_ok=True)
        out_path = os.path.join(out_dir, 'testcases_all.xlsx')

        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        HEADER_FONT = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        SMOKE_FILL = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
        P0_FILL = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
        P1_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
        LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
        THIN = Border(*[Side(style='thin', color='BFBFBF')] * 4)

        # 总览
        ws = wb.create_sheet('总览', 0)
        ws.append(['需求梳理与用例生成 - 总览'])
        ws['A1'].font = Font(name='微软雅黑', size=14, bold=True)
        p0 = sum(1 for c in testcases if c.get('priority') == 'P0')
        p1 = sum(1 for c in testcases if c.get('priority') == 'P1')
        p2 = sum(1 for c in testcases if c.get('priority') == 'P2')
        smoke_cases = smoke_json.get('smoke_testcases', []) if isinstance(smoke_json, dict) else []
        smoke_cnt = len(smoke_cases)
        stats = [
            ('模块数', len({c.get('module') for c in testcases})),
            ('用例总数', len(testcases)),
            ('P0', p0), ('P1', p1), ('P2', p2),
            ('冒烟用例数', smoke_cnt),
            ('生成时间', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ]
        ws.append([])
        ws.append(['统计项', '数值'])
        for c in (1, 2):
            ws.cell(row=3, column=c).font = HEADER_FONT
            ws.cell(row=3, column=c).fill = HEADER_FILL
        for label, val in stats:
            ws.append([label, val])
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15

        # 用例 sheet（按模块）
        modules = {}
        for c in testcases:
            modules.setdefault(c.get('module', '未命名'), []).append(c)
        cols = [('ID', 14), ('模块', 15), ('标题', 35), ('前置条件', 25), ('测试步骤', 50),
                ('预期结果', 30), ('优先级', 8), ('类型', 10), ('关联需求', 15), ('是否冒烟', 10), ('备注', 20)]
        for mod_name, cases in modules.items():
            ws2 = wb.create_sheet(mod_name[:31])
            ws2.append([c[0] for c in cols])
            for ci, _ in enumerate(cols, 1):
                cell = ws2.cell(row=1, column=ci)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = CENTER
                cell.border = THIN
            for c in cases:
                steps = c.get('steps', [])
                steps_txt = '\n'.join(f'{i+1}. {s}' for i, s in enumerate(steps)) if isinstance(steps, list) else str(steps)
                ws2.append([
                    c.get('id', ''), c.get('module', ''), c.get('title', ''),
                    c.get('precondition', ''), steps_txt, c.get('expected', ''),
                    c.get('priority', ''), c.get('type', ''), c.get('requirement_ref', ''),
                    '是' if c.get('is_smoke') else '否', c.get('notes', ''),
                ])
                r = ws2.max_row
                for ci in range(1, len(cols) + 1):
                    cell = ws2.cell(row=r, column=ci)
                    cell.alignment = LEFT
                    cell.border = THIN
                    cell.font = Font(name='微软雅黑', size=10)
                    if ci == 7 and c.get('priority') in ('P0', 'P1'):
                        cell.fill = P0_FILL if c.get('priority') == 'P0' else P1_FILL
                        cell.alignment = CENTER
                    if ci == 10 and c.get('is_smoke'):
                        cell.fill = SMOKE_FILL
                        cell.alignment = CENTER
                ws2.row_dimensions[r].height = 60
            for ci, (_, w) in enumerate(cols, 1):
                ws2.column_dimensions[get_column_letter(ci)].width = w
            ws2.freeze_panes = 'A2'

        # 冒烟 sheet
        if smoke_cases:
            ws3 = wb.create_sheet('冒烟用例')
            shead = ['序号', '用例ID', '所属模块', '冒烟标题', '前置条件', '测试步骤', '通过标准', '执行结果', '执行人', '备注']
            ws3.append(shead)
            for ci, h in enumerate(shead, 1):
                cell = ws3.cell(row=1, column=ci)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = CENTER
                cell.border = THIN
            for i, sc in enumerate(smoke_cases, 1):
                steps = sc.get('test_steps') or sc.get('steps') or []
                steps_txt = '\n'.join(f'{j+1}. {s}' for j, s in enumerate(steps)) if isinstance(steps, list) else str(steps)
                ws3.append([i, sc.get('id', ''), sc.get('module', ''), sc.get('title', ''),
                            sc.get('precondition', ''), steps_txt,
                            sc.get('pass_criteria') or sc.get('expected') or '', '', '', sc.get('notes', '')])
                for ci in range(1, len(shead) + 1):
                    cell = ws3.cell(row=ws3.max_row, column=ci)
                    cell.alignment = LEFT
                    cell.border = THIN
                    cell.font = Font(name='微软雅黑', size=10)
            ws3.freeze_panes = 'A2'

        wb.save(out_path)
        logger.info(f'Excel 导出成功: {out_path}')
        return out_path

    # ---------- 阶段4：冒烟执行决策（前端驱动）----------
    def record_smoke_decision(self, decision: str, reject_email: str = ''):
        self.prototype.stage4_decision = decision  # passed / rejected
        if reject_email:
            self.prototype.clarification_log = (self.prototype.clarification_log or '') + '\n\n' + reject_email
        self.prototype.current_stage = 4
        self.prototype.save()
