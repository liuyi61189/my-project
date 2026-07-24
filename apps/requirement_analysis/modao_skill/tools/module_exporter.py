#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
module_exporter.py
按模块导出测试用例到 Excel（每个模块一个 sheet）

使用：
  python module_exporter.py --input output/testcases/ --output output/testcases_all.xlsx \
                            --smoke-file output/testcases/smoke_testcases.md

依赖：
  pip install openpyxl
"""

import argparse
import os
import re
import sys
from pathlib import Path
from typing import List, Dict, Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# 表格样式定义
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") if OPENPYXL_AVAILABLE else None
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF") if OPENPYXL_AVAILABLE else None
SMOKE_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid") if OPENPYXL_AVAILABLE else None
P0_FILL = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid") if OPENPYXL_AVAILABLE else None
P1_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") if OPENPYXL_AVAILABLE else None
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True) if OPENPYXL_AVAILABLE else None
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True) if OPENPYXL_AVAILABLE else None
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
) if OPENPYXL_AVAILABLE else None


COLUMNS = [
    ("ID", 12),
    ("模块", 15),
    ("标题", 35),
    ("前置条件", 25),
    ("测试步骤", 50),
    ("预期结果", 30),
    ("优先级", 8),
    ("类型", 10),
    ("关联需求", 15),
    ("是否冒烟", 10),
    ("备注", 20),
]


def parse_testcase_md(md_path: Path) -> List[Dict[str, Any]]:
    """
    解析单个模块的测试用例 Markdown 文件，提取结构化测试用例列表。

    解析规则（兼容 testcase_module.md 模板）：
    - 表头行：| 用例ID | 模块 | 标题 | 前置条件 | 步骤 | 预期结果 | 优先级 | 关联需求 | 用例类型 | 备注 |
    - 冒烟标记：标题前带 [冒烟] 前缀的视为冒烟用例
    """
    if not md_path.exists():
        return []

    content = md_path.read_text(encoding="utf-8")
    testcases = []

    # 找到所有 markdown 表格块
    table_pattern = re.compile(
        r"^\|(.+)\|\s*$\n^\|[\s\-:|]+\|\s*$\n((?:^\|.+\|\s*$\n?)+)",
        re.MULTILINE,
    )

    for match in table_pattern.finditer(content):
        header_line = match.group(1)
        body_lines = match.group(2).strip().split("\n")

        headers = [h.strip() for h in header_line.split("|")]

        # 仅处理包含"用例ID"或"标题"列的表
        if not any(h in ("用例ID", "标题", "冒烟ID") for h in headers):
            continue

        for line in body_lines:
            cells = [c.strip() for c in line.strip("|").split("|")]
            if len(cells) != len(headers):
                continue
            row = dict(zip(headers, cells))
            testcases.append(row)

    return testcases


def parse_smoke_md(md_path: Path) -> List[Dict[str, Any]]:
    """
    解析冒烟用例 Markdown 文件。
    """
    if not md_path.exists():
        return []

    content = md_path.read_text(encoding="utf-8")
    smoke_cases = []

    # 冒烟表头：| 序号 | 用例ID | 所属模块 | 冒烟标题 | ...
    table_pattern = re.compile(
        r"^\|(.+)\|\s*$\n^\|[\s\-:|]+\|\s*$\n((?:^\|.+\|\s*$\n?)+)",
        re.MULTILINE,
    )

    for match in table_pattern.finditer(content):
        header_line = match.group(1)
        body_lines = match.group(2).strip().split("\n")

        headers = [h.strip() for h in header_line.split("|")]
        if "冒烟标题" not in headers and "标题" not in headers:
            continue

        for line in body_lines:
            cells = [c.strip() for c in line.strip("|").split("|")]
            if len(cells) != len(headers):
                continue
            row = dict(zip(headers, cells))
            smoke_cases.append(row)

    return smoke_cases


def get_priority_fill(priority: str):
    """根据优先级返回对应填充色"""
    if not OPENPYXL_AVAILABLE:
        return None
    if priority == "P0":
        return P0_FILL
    elif priority == "P1":
        return P1_FILL
    return None


def write_overview_sheet(wb: Workbook, modules_data: Dict[str, List[Dict]], smoke_cases: List[Dict]):
    """写入「总览」sheet"""
    ws = wb.create_sheet("总览", 0)
    ws.append(["需求梳理与用例生成 - 总览"])
    ws["A1"].font = Font(name="微软雅黑", size=14, bold=True)
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 30

    ws.append([])
    ws.append(["统计项", "数值"])
    ws["A3"].font = HEADER_FONT
    ws["B3"].font = HEADER_FONT
    ws["A3"].fill = HEADER_FILL
    ws["B3"].fill = HEADER_FILL
    ws["A3"].alignment = CENTER
    ws["B3"].alignment = CENTER

    total = sum(len(tcs) for tcs in modules_data.values())
    p0 = sum(1 for tcs in modules_data.values() for tc in tcs if tc.get("优先级") == "P0")
    p1 = sum(1 for tcs in modules_data.values() for tc in tcs if tc.get("优先级") == "P1")
    p2 = sum(1 for tcs in modules_data.values() for tc in tcs if tc.get("优先级") == "P2")
    smoke_count = len(smoke_cases)
    smoke_ratio = (smoke_count / p0 * 100) if p0 > 0 else 0

    stats = [
        ("模块数", len(modules_data)),
        ("用例总数", total),
        ("P0 用例数", p0),
        ("P1 用例数", p1),
        ("P2 用例数", p2),
        ("冒烟用例数", smoke_count),
        ("冒烟用例占 P0 比例", f"{smoke_ratio:.1f}%"),
        ("生成时间", __import__("datetime").datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for label, val in stats:
        ws.append([label, val])
        a_cell = ws.cell(row=ws.max_row, column=1)
        b_cell = ws.cell(row=ws.max_row, column=2)
        a_cell.font = Font(name="微软雅黑", size=11, bold=True)
        b_cell.font = Font(name="微软雅黑", size=11)

    # 模块级汇总
    ws.append([])
    ws.append(["模块", "用例数", "P0", "P1", "P2", "是否含冒烟"])
    for col in range(1, 7):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    for module_name, tcs in modules_data.items():
        module_p0 = sum(1 for tc in tcs if tc.get("优先级") == "P0")
        module_p1 = sum(1 for tc in tcs if tc.get("优先级") == "P1")
        module_p2 = sum(1 for tc in tcs if tc.get("优先级") == "P2")
        has_smoke = any(tc.get("是否冒烟") in ("是", "True", "true", "✓") for tc in tcs)
        ws.append([module_name, len(tcs), module_p0, module_p1, module_p2, "是" if has_smoke else "否"])
        for col in range(1, 7):
            ws.cell(row=ws.max_row, column=col).alignment = CENTER
            ws.cell(row=ws.max_row, column=col).border = THIN_BORDER

    # 列宽
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    for col in "CDEF":
        ws.column_dimensions[col].width = 12


def write_module_sheet(wb: Workbook, module_name: str, testcases: List[Dict]):
    """为单个模块写入 sheet"""
    # Sheet 名最长 31 字符
    sheet_name = module_name[:31] if len(module_name) > 31 else module_name
    ws = wb.create_sheet(sheet_name)

    # 写入表头
    headers = [col[0] for col in COLUMNS]
    ws.append(headers)
    for col_idx, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # 列宽
    for col_idx, (_, width) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 写入数据
    for tc in testcases:
        row_data = [
            tc.get("用例ID", ""),
            tc.get("模块", module_name),
            tc.get("标题", ""),
            tc.get("前置条件", ""),
            tc.get("步骤", "").replace("<br>", "\n").replace("1.", "1.").replace("2.", "\n2."),
            tc.get("预期结果", ""),
            tc.get("优先级", ""),
            tc.get("用例类型", ""),
            tc.get("关联需求", ""),
            tc.get("是否冒烟", "否"),
            tc.get("备注", ""),
        ]
        ws.append(row_data)
        current_row = ws.max_row
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.alignment = LEFT
            cell.border = THIN_BORDER
            cell.font = Font(name="微软雅黑", size=10)

            # 优先级填充
            priority = tc.get("优先级", "")
            if priority in ("P0", "P1") and col_idx == 7:
                cell.fill = get_priority_fill(priority)
                cell.alignment = CENTER

            # 冒烟标记填充
            if col_idx == 10 and tc.get("是否冒烟") in ("是", "True", "true", "✓"):
                cell.fill = SMOKE_FILL
                cell.alignment = CENTER

        ws.row_dimensions[current_row].height = 60

    # 冻结首行
    ws.freeze_panes = "A2"


def write_smoke_sheet(wb: Workbook, smoke_cases: List[Dict]):
    """写入「冒烟用例」sheet"""
    ws = wb.create_sheet("冒烟用例")
    headers = ["序号", "用例ID", "所属模块", "冒烟标题", "前置条件", "测试步骤", "通过标准", "执行结果", "执行人", "备注"]
    ws.append(headers)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    widths = [6, 15, 12, 30, 20, 40, 25, 10, 10, 15]
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    for idx, sc in enumerate(smoke_cases, 1):
        ws.append([
            idx,
            sc.get("用例ID", ""),
            sc.get("所属模块", ""),
            sc.get("冒烟标题", ""),
            sc.get("前置条件", ""),
            sc.get("测试步骤", ""),
            sc.get("通过标准", ""),
            "",  # 实际结果 - 留空给执行人填写
            "",  # 执行人
            sc.get("备注", ""),
        ])
        current_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.alignment = LEFT
            cell.border = THIN_BORDER
            cell.font = Font(name="微软雅黑", size=10)
        ws.row_dimensions[current_row].height = 50

    ws.freeze_panes = "A2"


def write_statistics_sheet(wb: Workbook, modules_data: Dict, smoke_cases: List):
    """写入「统计」sheet"""
    ws = wb.create_sheet("统计")
    ws.append(["需求梳理与用例生成 - 统计报表"])
    ws["A1"].font = Font(name="微软雅黑", size=14, bold=True)
    ws.merge_cells("A1:E1")

    ws.append([])
    ws.append(["优先级分布"])
    ws["A3"].font = Font(name="微软雅黑", size=12, bold=True)
    ws.append(["优先级", "数量", "占比"])
    for col in range(1, 4):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    total = sum(len(tcs) for tcs in modules_data.values())
    p0 = sum(1 for tcs in modules_data.values() for tc in tcs if tc.get("优先级") == "P0")
    p1 = sum(1 for tcs in modules_data.values() for tc in tcs if tc.get("优先级") == "P1")
    p2 = sum(1 for tcs in modules_data.values() for tc in tcs if tc.get("优先级") == "P2")

    for label, cnt in [("P0", p0), ("P1", p1), ("P2", p2)]:
        ratio = f"{(cnt / total * 100):.1f}%" if total > 0 else "0%"
        ws.append([label, cnt, ratio])

    ws.append([])
    ws.append(["类型分布"])
    type_label_cell = ws.cell(row=ws.max_row, column=1)
    type_label_cell.font = Font(name="微软雅黑", size=12, bold=True)
    ws.append(["类型", "数量"])
    for col in range(1, 3):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    type_count: Dict[str, int] = {}
    for tcs in modules_data.values():
        for tc in tcs:
            t = tc.get("用例类型", "未分类")
            type_count[t] = type_count.get(t, 0) + 1
    for t, c in type_count.items():
        ws.append([t, c])

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15


def export(input_dir: str, output_path: str, smoke_file: str = None) -> int:
    """
    主导出函数
    Returns: 0 成功 / 1 失败
    """
    if not OPENPYXL_AVAILABLE:
        print("⚠️ openpyxl 不可用，无法生成 Excel。请执行: pip install openpyxl")
        return 1

    input_path = Path(input_dir)
    if not input_path.exists():
        print(f"❌ 输入目录不存在: {input_dir}")
        return 1

    # 解析所有模块的用例
    md_files = sorted(input_path.glob("*.md"))
    # 排除冒烟用例文件
    md_files = [f for f in md_files if "smoke" not in f.name.lower() and "冒烟" not in f.name]

    modules_data: Dict[str, List[Dict]] = {}
    for md_file in md_files:
        module_name = md_file.stem
        testcases = parse_testcase_md(md_file)
        if testcases:
            modules_data[module_name] = testcases

    if not modules_data:
        print("⚠️ 未解析到任何测试用例")
        return 1

    # 解析冒烟用例
    smoke_cases: List[Dict] = []
    if smoke_file and Path(smoke_file).exists():
        smoke_cases = parse_smoke_md(Path(smoke_file))

    # 创建工作簿
    wb = Workbook()
    # 删除默认 Sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # 写入各 sheet
    write_overview_sheet(wb, modules_data, smoke_cases)
    for module_name, testcases in modules_data.items():
        write_module_sheet(wb, module_name, testcases)
    if smoke_cases:
        write_smoke_sheet(wb, smoke_cases)
    write_statistics_sheet(wb, modules_data, smoke_cases)

    # 保存
    output_path_obj = Path(output_path)
    output_path_obj.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path_obj))

    # 输出统计
    total = sum(len(tcs) for tcs in modules_data.values())
    p0 = sum(1 for tcs in modules_data.values() for tc in tcs if tc.get("优先级") == "P0")
    smoke_count = len(smoke_cases)
    smoke_ratio = (smoke_count / p0 * 100) if p0 > 0 else 0

    print(f"✅ Excel 导出成功: {output_path}")
    print(f"📊 统计：")
    print(f"   - 模块数：{len(modules_data)}")
    print(f"   - 用例总数：{total}")
    print(f"   - 冒烟用例：{smoke_count}（占 P0: {smoke_ratio:.1f}%）")

    return 0


def main():
    parser = argparse.ArgumentParser(description="按模块导出测试用例到 Excel")
    parser.add_argument("--input", "-i", required=True, help="输入目录（包含各模块的 .md 用例文件）")
    parser.add_argument("--output", "-o", required=True, help="输出 Excel 文件路径")
    parser.add_argument("--smoke-file", "-s", default=None, help="冒烟用例 Markdown 文件路径（可选）")

    args = parser.parse_args()
    exit_code = export(args.input, args.output, args.smoke_file)
    sys.exit(exit_code)


if __name__ == "__main__":
    main()
